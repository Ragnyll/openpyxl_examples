import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rocket import Rocket

def generate_square_data(sheet):
    counter = 0
    for row in range(2, 10):
        for col in range(2,10):
            cell = sheet.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
            cell.value = counter
            counter += 1

def create_rocket_data(rocket: Rocket):
    # opens with 1 default sheet
    wb = Workbook()
    wb.active.title = '{} RocketStats'.format(rocket.name)
    rocket.write_to_sheet(wb.active)

    # Create 2 data sheets
    for i in range(2):
        ws = wb.create_sheet('DataSheet{}'.format(i))
        generate_square_data(ws)


    wb.save('RocketStats.xlsx')
    return


if __name__ == '__main__':
    if sys.argv[1] == 'generate':
        myrocket = Rocket('funky', '1000', '349')
        create_rocket_data(myrocket)
    elif sys.argv[1] == 'derive':
        print('gotta implement this')

