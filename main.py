import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from element import Element
import random


def generate_square_data(sheet):
    counter = 0
    for row in range(2, 10):
        skipped_cell = False
        for col in range(2, 6):
            if not skipped_cell and random.randrange(0, 4, 1) == 0:
                skipped_cell = True
                continue
            cell = sheet.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
            cell.value = counter
            counter += 1


def create_element_data(element: Element):
    # opens with 1 default sheet
    wb = Workbook()
    wb.active.title = '{} ElementStats'.format(element.name)
    element.write_to_sheet(wb.active)

    # Create 2 data sheets
    for i in range(2):
        ws = wb.create_sheet('DataSheet{}'.format(i))
        generate_square_data(ws)


    wb.save('Element.xlsx')
    return


if __name__ == '__main__':
    if sys.argv[1] == 'generate':
        hydrogen = Element('Hydrogen', '8.675309', '1')
        create_element_data(hydrogen)
    elif sys.argv[1] == 'derive':
        print('gotta implement this')

